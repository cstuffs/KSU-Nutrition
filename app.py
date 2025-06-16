from flask import Flask, render_template, request, redirect, url_for, session, send_file
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from collections import OrderedDict
from models import Order, db
import json
import os
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your-secret-key'

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Dummy user class using team name as ID
class User(UserMixin):
    def __init__(self, username, team):
        self.id = username
        self.team = team

@login_manager.user_loader
def load_user(user_id):
    # Get team from session or fallback to dummy
    team = session.get("team", "Unknown Team")
    return User(user_id, team)

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)

# Constants
EXCEL_DIR = 'user_orders'
os.makedirs(EXCEL_DIR, exist_ok=True)
USERS_FILE = 'users.json'
BUDGET_FILE = 'budgets.json'

# Load menu
with open('structured_menu.json', 'r') as f:
    full_menu = json.load(f, object_pairs_hook=OrderedDict)

# === Helper Functions ===

def load_users():
    with open(USERS_FILE, 'r') as f:
        return json.load(f, object_pairs_hook=OrderedDict)

def build_member_to_team(users):
    mtt = {}
    for team, members in users.items():
        for m in members:
            m = m.strip()
            if m:
                mtt[m] = team
    return mtt

users = load_users()
member_to_team = build_member_to_team(users)

def calculate_total_spent_for_team(team_name):
    total_spent = 0.0

    # Load menu for price lookup
    with open("structured_menu.json", "r") as f:
        menu = json.load(f)

    price_lookup = {
        f"{item}|||{opt['name']}": opt["price"]
        for group in menu.values()
        for item, options in group.items()
        for opt in options
    }

    users = load_users()
    member_to_team = build_member_to_team(users)

    for filename in os.listdir(EXCEL_DIR):
        if not filename.endswith(".xlsx"):
            continue

        filepath = os.path.join(EXCEL_DIR, filename)
        wb = load_workbook(filepath, data_only=True)

        if "Yearly Orders" not in wb.sheetnames:
            continue

        sheet = wb["Yearly Orders"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                _, _, member_name, item_name, option_name, qty = row
                if member_to_team.get(member_name, "") == team_name:
                    key = f"{item_name}|||{option_name}"
                    price = price_lookup.get(key, 0.0)
                    total_spent += float(qty) * price
            except:
                continue

    return total_spent

def load_budgets():
    if not os.path.exists(BUDGET_FILE):
        return {}
    with open(BUDGET_FILE, 'r') as f:
        return json.load(f)

def save_budgets(budgets):
    with open(BUDGET_FILE, 'w') as f:
        json.dump(budgets, f, indent=2)

def get_week_range():
    today = datetime.now()
    start = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end = start + timedelta(days=6)
    return start, end

# === Database save orders===

def save_user_order(member_name, order_datetime, items):
    team_name = session.get("team", "Unknown Team")

    for item in items:
        order = Order(
            team=team_name,
            member=member_name,
            date=order_datetime.date(),
            time=order_datetime.time(),
            item_name=item["name"],
            option=item.get("option", ""),
            quantity=item["quantity"],
            price=item["price"]  # ✅ Ensure price is included in each item
        )
        db.session.add(order)

    db.session.commit()

@app.route('/')
def home():
    return redirect(url_for('submit_order'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        raw_team = request.form.get('team_name', '').strip()
        raw_name = request.form.get('member_name', '').strip()

        team_input = raw_team.lower()
        member_input = raw_name.lower()

        users = load_users()
        users_normalized = {
            team.lower(): [m.strip().lower() for m in members]
            for team, members in users.items()
        }

        if team_input not in users_normalized:
            return f"Team '{raw_team}' not found.", 403

        if member_input not in users_normalized[team_input]:
            return f"User '{raw_name}' not found on team '{raw_team}'.", 403

        # === Scott Trausch is full admin ===
        if raw_team == "KSU Football" and raw_name == "Scott Trausch":
            login_user(User("admin", "KSU Football"))
            session['team'] = "KSU Football"
            session['member_name'] = "Scott Trausch"
            session['admin_as_football'] = False  # full admin
            return redirect(url_for('admin_dashboard'))

        # === Other Football users get limited dashboard (no edit buttons) ===
        if raw_team == "KSU Football":
            login_user(User(raw_name, raw_team))
            session['team'] = "KSU Football"
            session['member_name'] = raw_name
            session['admin_as_football'] = True  # limited admin
            return redirect(url_for('admin_dashboard'))

        # === All other teams ===
        login_user(User(raw_name, raw_team))
        session['team'] = raw_team
        session['member_name'] = raw_name
        session['admin_as_football'] = False
        return redirect(url_for('submit_order'))

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    return redirect(url_for('login'))

@app.route('/order', methods=['GET'])
@login_required
def submit_order():
    # Allow admin to access the order page if they are acting as KSU Football
    if current_user.id == 'admin' and not session.get('admin_as_football'):
        return redirect(url_for('admin_dashboard'))

    if request.args.get("new") == "1":
        session.pop("last_order_form", None)

    form_data = session.get('last_order_form', {})

    # ✅ Get team name from session
    team_name = session.get("team", "unknown_team")

    # ✅ Load full team budget
    budgets = load_budgets()
    team_budget = budgets.get(team_name, 100.00)  # fallback if missing

    # ✅ Calculate total spent from submitted orders (Excel-based)
    total_spent = calculate_total_spent_for_team(team_name)
    remaining_budget = team_budget - total_spent

    # ✅ Week range string
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    # ✅ Load menu structure
    with open('structured_menu.json', 'r') as f:
        grouped_menu = json.load(f, object_pairs_hook=OrderedDict)

    return render_template("order.html",
                           current_user=current_user,
                           session=session,
                           grouped_menu=grouped_menu,
                           user_budget=team_budget,
                           remaining_budget=remaining_budget,
                           week_range=week_range_str,
                           form_data=form_data)

@app.route('/add_to_order', methods=['POST'])
@login_required
def add_to_order():
    form_data = request.form.to_dict()

    print("🧪 Incoming form_data:", form_data)  # ✅ Move it here

    # Only save fields with positive quantity
    cleaned_form = {}

    for key in form_data:
        if key.startswith("qty_") and form_data[key].isdigit() and int(form_data[key]) > 0:
            cleaned_form[key] = form_data[key]

        elif key.startswith("meta_"):
            suffix = key[5:]
            qty_key = "qty_" + suffix
            if qty_key in form_data and form_data[qty_key].isdigit() and int(form_data[qty_key]) > 0:
                cleaned_form[key] = form_data[key]

    session['last_order_form'] = cleaned_form

    if form_data.get("action") == "review":
        return redirect(url_for('review_order'))
    else:
        return redirect(url_for('submit_order'))

@app.route('/order/edit', methods=['POST'])
@login_required
def order_form_edit():
    if current_user.id == 'admin' and not session.get("admin_as_football"):
        return redirect(url_for('admin_dashboard'))

    form_data = request.form
    selected_items = []

    for key in form_data:
        if key.startswith("meta_"):
            qty_key = "qty_" + key[5:]
            try:
                quantity = int(form_data.get(qty_key, 0))
                if quantity > 0:
                    item_name, option_name, price = form_data[key].split("|||")
                    selected_items.append({
                        "name": item_name,
                        "option": option_name,
                        "price": float(price),
                        "quantity": quantity,
                        "meta_key": key,
                        "qty_key": qty_key
                    })
            except Exception as e:
                print(f"⚠️ Error parsing {key}: {e}")
                continue

    return render_template("order_edit.html", selected_items=selected_items)

@app.route('/order/review', methods=['GET', 'POST'])
@login_required
def review_order():
    budgets = load_budgets()
    team_name = session.get("team")
    team_budget = budgets.get(team_name, 100.00)

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    form_data = session.get('last_order_form', {})
    session['last_order_form'] = form_data  # Save for "Return to Order Page"

    items = []
    total = 0.0

    # Iterate over all submitted keys
    for key in form_data:
        if key.startswith("meta_"):
            suffix = key[5:]  # e.g., Gatorade_Shake_Chocolate
            qty_key = "qty_" + suffix
            qty_str = form_data.get(qty_key, "0").strip()

            if qty_str.isdigit() and int(qty_str) > 0:
                quantity = int(qty_str)
                try:
                    item_name, option_name, price = form_data[key].split("|||")
                    price = float(price)
                    subtotal = price * quantity
                    total += subtotal
                    items.append({
                        "name": item_name,
                        "option": option_name,
                        "price": price,
                        "quantity": quantity,
                        "subtotal": subtotal
                    })
                except Exception as e:
                    print(f"⚠️ Error parsing key {key}: {e}")
                    continue

    return render_template("order_review.html",
                           items=items,
                           total=total,
                           user_budget=team_budget,
                           remaining_budget=team_budget - total,
                           week_range=week_range_str,
                           form_data=form_data)

@app.route('/order/submit', methods=['POST'])
@login_required
def finalize_order():
    form_data = session.get('last_order_form', {})

    items = []
    total = 0.0

    for key in form_data:
        if key.startswith("meta_"):
            suffix = key[5:]
            qty_key = "qty_" + suffix
            qty_str = form_data.get(qty_key, "0").strip()
            if qty_str.isdigit() and int(qty_str) > 0:
                item_name, option_name, price = form_data[key].split("|||")
                price = float(price)
                quantity = int(qty_str)
                total += price * quantity
                items.append({
                    "name": item_name,
                    "option": option_name,
                    "price": price,
                    "quantity": quantity
                })

    # ✅ Save order if any items exist
    if items:
        save_user_order(session.get("member_name"), datetime.now(), items)

    # ✅ Load team budget and recalculate total spent from Excel
    team_name = session.get("team", "unknown_team")
    budgets = load_budgets()
    team_budget = budgets.get(team_name, 100.00)
    total_spent = calculate_total_spent_for_team(team_name)
    remaining_budget = team_budget - total_spent  # Optional: for logging/debugging if needed

    # ✅ Clear form after submission
    session.pop('last_order_form', None)

    return redirect(url_for('submit_order'))

from sqlalchemy import and_

@app.route('/admin/produce_hyvee')
@login_required
def admin_produce_hyvee():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    # Load Produce and Hyvee items from structured_menu
    with open("structured_menu.json", "r") as f:
        menu = json.load(f)

    valid_items = set(menu.get("Produce", {}).keys()) | set(menu.get("Hyvee", {}).keys())

    # Query the DB for this week's orders
    results = Order.query.filter(
        and_(
            Order.date >= start_of_week.date(),
            Order.date <= end_of_week.date(),
            Order.item_name.in_(valid_items)
        )
    ).order_by(Order.date.desc()).all()

    # Format results
    filtered_orders = [
        (order.date.strftime("%-m/%-d/%y"), order.team, order.item_name, order.quantity)
        for order in results
    ]

    week_range = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"
    return render_template("admin_produce_hyvee.html", orders=filtered_orders, week_range=week_range)

@app.route('/admin')
@login_required
def admin_dashboard():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    users = load_users()
    all_teams = list(users.keys())

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    return render_template('admin_dashboard.html', teams=all_teams, week_range=week_range_str)

@app.route('/admin/football_order')
@login_required
def admin_football_order():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    # Only admin can simulate football ordering (do NOT overwrite session info)
    session['admin_as_football'] = True
    return redirect(url_for('submit_order'))

@app.route('/admin/team/<team_name>')
@login_required
def view_team_orders(team_name):
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)
    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    results = Order.query.filter(
        Order.team == team_name,
        Order.date >= start_of_week.date(),
        Order.date <= end_of_week.date()
    ).order_by(Order.date, Order.time).all()

    # Weekly orders grouped by member
    orders_by_member = {}
    all_totals = {}

    for o in results:
        item = f"{o.item_name} - {o.option}".strip(" -")
        subtotal = o.quantity * o.price

        if o.member not in orders_by_member:
            orders_by_member[o.member] = {"orders": [], "total": 0.0}

        orders_by_member[o.member]["orders"].append({
            "date": o.date.strftime("%Y-%m-%d"),
            "time": o.time.strftime("%I:%M %p"),
            "item": item,
            "quantity": o.quantity,
            "price": f"${o.price:.2f}",
            "subtotal": f"${subtotal:.2f}"
        })

        orders_by_member[o.member]["total"] += subtotal

        if item not in all_totals:
            all_totals[item] = {"qty": 0, "total_cost": 0.0}

        all_totals[item]["qty"] += o.quantity
        all_totals[item]["total_cost"] += subtotal

    total_team_cost = sum([v["total_cost"] for v in all_totals.values()])
    team_budget = load_budgets().get(team_name, 100.00)
    remaining_budget = team_budget - total_team_cost

    return render_template("team_orders.html",
                           team_name=team_name,
                           week_range=week_range_str,
                           weekly_orders_by_member=orders_by_member,
                           total_orders=all_totals,
                           total_cost=total_team_cost,
                           user_budget=team_budget,
                           remaining_budget=remaining_budget)

from openpyxl import Workbook
from io import BytesIO
from flask import send_file

from openpyxl import Workbook
from io import BytesIO
from flask import send_file

@app.route('/admin/produce_hyvee/export')
@login_required
def export_produce_hyvee_excel():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    with open("structured_menu.json", "r") as f:
        menu = json.load(f)

    valid_items = set(menu.get("Produce", {}).keys()) | set(menu.get("Hyvee", {}).keys())

    results = Order.query.filter(
        and_(
            Order.date >= start_of_week.date(),
            Order.date <= end_of_week.date(),
            Order.item_name.in_(valid_items)
        )
    ).order_by(Order.date.asc(), Order.team.asc()).all()

    # Create workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Produce & Hyvee"
    ws_out.append(["Date", "Team", "Item", "Quantity"])

    for order in results:
        ws_out.append([
            order.date.strftime("%Y-%m-%d"),
            order.team,
            order.item_name,
            order.quantity
        ])

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    filename = f"Produce_Hyvee_Orders_{start_of_week.strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

from io import BytesIO

@app.route('/admin/weekly_summary/export')
@login_required
def export_weekly_summary_excel():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    results = Order.query.filter(
        and_(
            Order.date >= start_of_week.date(),
            Order.date <= end_of_week.date()
        )
    ).order_by(Order.date.asc(), Order.time.asc()).all()

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Weekly Summary"
    ws_out.append(["Date", "Team", "Item", "Quantity"])

    for order in results:
        item_full = f"{order.item_name} - {order.option}".strip(" -")
        ws_out.append([
            order.date.strftime("%Y-%m-%d"),
            order.team,
            item_full,
            order.quantity
        ])

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)

    filename = f"Full_Weekly_Orders_{start_of_week.strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     download_name=filename,
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/admin/weekly_summary')
@login_required
def admin_weekly_summary():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 1) if today.weekday() != 6 else today
    end_of_week = start_of_week + timedelta(days=6)

    week_range_str = f"{start_of_week.strftime('%-m/%-d/%y')} - {end_of_week.strftime('%-m/%-d/%y')}"

    results = Order.query.filter(
        and_(
            Order.date >= start_of_week.date(),
            Order.date <= end_of_week.date()
        )
    ).order_by(Order.date.desc(), Order.time.desc()).all()

    all_orders = []
    for o in results:
        item_full = f"{o.item_name} - {o.option}".strip(" -")
        all_orders.append({
            "date": o.date.strftime("%-m/%-d/%y"),
            "team": o.team,
            "item": item_full,
            "quantity": o.quantity
        })

    return render_template("weekly_summary.html", week_range=week_range_str, orders=all_orders)

@app.route('/admin/weekly_totals')
@login_required
def weekly_totals():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    def get_week_number(date):
        week1_start = datetime(2025, 1, 1)
        week1_start -= timedelta(days=week1_start.weekday() + 1) if week1_start.weekday() != 6 else timedelta(0)
        delta = date - week1_start
        return (delta.days // 7) + 1

    results = Order.query.all()

    totals_by_week = {week: {} for week in range(1, 53)}
    yearly_totals_by_week = {2025: {w: {} for w in range(1, 53)}, 2024: {w: {} for w in range(1, 53)}}

    for o in results:
        year = o.date.year
        week_num = get_week_number(o.date)
        subtotal = o.price * o.quantity

        if week_num < 1 or week_num > 52:
            continue

        if year not in yearly_totals_by_week:
            yearly_totals_by_week[year] = {w: {} for w in range(1, 53)}

        yearly_totals_by_week[year][week_num][o.team] = yearly_totals_by_week[year][week_num].get(o.team, 0.0) + subtotal

        if year == 2025:
            totals_by_week[week_num][o.team] = totals_by_week[week_num].get(o.team, 0.0) + subtotal

    users = load_users()

    return render_template("weekly_totals.html",
                           totals_by_week=totals_by_week,
                           yearly_totals_by_week=yearly_totals_by_week,
                           users=users,
                           datetime=datetime,
                           timedelta=timedelta)

@app.route('/admin/all_orders')
@login_required
def all_orders():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    results = Order.query.order_by(Order.date.desc(), Order.time.desc()).all()

    full_orders = []
    for o in results:
        order_date = o.date
        week_num = (order_date - datetime(2025, 1, 1)).days // 7 + 1

        full_orders.append({
            "date": o.date.strftime("%Y-%m-%d"),
            "time": o.time.strftime("%I:%M %p"),
            "week": week_num,
            "year": o.date.year,
            "team": o.team,
            "member": o.member,
            "item": f"{o.item_name} - {o.option}".strip(" -"),
            "quantity": o.quantity
        })

    return render_template("all_orders.html", orders=full_orders)

@app.route('/admin/budgets', methods=['GET', 'POST'])
@login_required
def manage_budgets():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    users = load_users()
    existing_budgets = load_budgets()
    updated_budgets = {}

    # Always include every team from users.json in the form
    for team in users.keys():
        updated_budgets[team] = existing_budgets.get(team, 100.00)  # default to $100 if missing

    if request.method == 'POST':
        for team in users.keys():
            new_value = request.form.get(team)
            if new_value:
                try:
                    updated_budgets[team] = float(new_value)
                except ValueError:
                    continue
        save_budgets(updated_budgets)
        return redirect(url_for('manage_budgets'))

    return render_template("manage_budgets.html", team_budgets=updated_budgets)

@app.route('/admin/edit_menu', methods=['GET', 'POST'])
@login_required
def edit_menu():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    menu_path = 'structured_menu.json'

    if request.method == 'POST':
        form = request.form
        updated_menu = OrderedDict()

        group_names = [key.split('[')[1].split(']')[0] for key in form.keys() if key.startswith("group_names[")]
        group_names = list(OrderedDict.fromkeys(group_names))

        for group in group_names:
            item_names = form.getlist(f'group_names[{group}][item_names][]')
            group_data = OrderedDict()
            for item_name in item_names:
                item_name = item_name.strip()
                if not item_name:
                    continue
                options = form.getlist(f'options[{item_name}][]')
                prices = form.getlist(f'prices[{item_name}][]')
                if not options or not prices or len(options) != len(prices):
                    continue
                item_options = []
                for opt, price_str in zip(options, prices):
                    opt = opt.strip()
                    try:
                        price = float(price_str)
                        if opt:
                            item_options.append({ "name": opt, "price": price })
                    except ValueError:
                        continue
                if item_options:
                    group_data[item_name] = item_options
            if group_data:
                updated_menu[group] = group_data

        with open(menu_path, 'w') as f:
            json.dump(updated_menu, f, indent=2)
        return redirect(url_for('edit_menu'))

    with open(menu_path, 'r') as f:
        grouped_menu = json.load(f, object_pairs_hook=OrderedDict)
    return render_template('edit_menu_fixed.html', grouped_menu=grouped_menu)

@app.route('/admin/edit_users', methods=['GET', 'POST'])
@login_required
def edit_users():
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    users_path = 'users.json'

    if request.method == 'POST':
        team_names = request.form.getlist('team_names[]')
        members_list = request.form.getlist('members[]')

        updated_users = OrderedDict()

        for team, members_raw in zip(team_names, members_list):
            members = [m.strip() for m in members_raw.splitlines() if m.strip()]
            updated_users[team.strip()] = members if members else [" "]

        with open(users_path, 'w') as f:
            json.dump(updated_users, f, indent=2)

        return redirect(url_for('admin_dashboard'))

    with open(users_path, 'r') as f:
        users = json.load(f, object_pairs_hook=OrderedDict)

    return render_template("edit_users.html", users=users)

@app.route('/admin/user/<user_name>')
@login_required
def view_user_file(user_name):
    if not (current_user.id == 'admin' or session.get('admin_as_football')):
        return "Access Denied", 403

    current_week = datetime.now().isocalendar()[1]

    weekly_orders = []
    total_orders = []

    results = Order.query.filter(Order.member == user_name).all()

    for o in results:
        order_week = o.date.isocalendar()[1]
        item_name = f"{o.item_name} - {o.option}".strip(" -")

        if order_week == current_week:
            weekly_orders.append({
                "date": o.date.strftime("%Y-%m-%d"),
                "item": item_name,
                "quantity": o.quantity
            })

        # accumulate total per item
        found = False
        for row in total_orders:
            if row["item"] == item_name:
                row["quantity"] += o.quantity
                found = True
                break
        if not found:
            total_orders.append({"item": item_name, "quantity": o.quantity})

    return render_template("user_orders.html",
                           user_name=user_name,
                           weekly_orders=weekly_orders,
                           total_orders=total_orders)

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # ← this creates the Order table
    app.run(debug=True)

if __name__ == "__main__":
    app.run(debug=True)
