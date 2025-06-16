import os
from openpyxl import load_workbook
from models import db, Order
from app import app
from datetime import datetime

EXCEL_DIR = "user_orders"

with app.app_context():
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]

    for file in files:
        filepath = os.path.join(EXCEL_DIR, file)
        wb = load_workbook(filepath)
        
        if "Yearly Orders" not in wb.sheetnames:
            continue
        
        ws = wb["Yearly Orders"]

        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_str, time_str, member, item_name, option, quantity = row

            try:
                order_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                order_time = datetime.strptime(time_str, "%H:%M:%S").time()
            except Exception as e:
                print(f"Skipping row due to bad date/time: {row} ({e})")
                continue

            # Guess team from filename
            team = file.replace("orders_", "").replace(".xlsx", "").strip()

            order = Order(
                team=team,
                member=member,
                date=order_date,
                time=order_time,
                item_name=item_name,
                option=option,
                quantity=int(quantity)
            )

            db.session.add(order)

    db.session.commit()
    print("✅ Migration complete.")
